"""End-to-end tests for the xlsx skill helper scripts.

Runs each script as a subprocess under LC_ALL=C to prove all text I/O
uses explicit UTF-8 rather than locale defaults. No network access.
"""
from __future__ import annotations

import csv
import json
import os
import shutil
import subprocess
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPTS = Path(__file__).resolve().parent.parent / "scripts"


def run(script, *args, expect_ok=True):
    env = dict(os.environ, LC_ALL="C", LANG="C")
    env.pop("PYTHONIOENCODING", None)
    proc = subprocess.run(
        [sys.executable, str(SCRIPTS / script), *map(str, args)],
        capture_output=True, text=True, env=env, encoding="utf-8")
    if expect_ok:
        assert proc.returncode == 0, f"{script} failed: {proc.stderr}"
    return proc


SPEC = {
    "full_calc_on_load": True,
    "sheets": [
        {
            "name": "Data",
            "rows": [
                [
                    {"value": "Region", "bold": True, "fill": "DDEBF7",
                     "border": "thin", "align": "center", "valign": "center"},
                    {"value": "Sales", "bold": True, "fill": "DDEBF7"},
                    {"value": "Growth", "bold": True},
                    {"value": "Audited", "bold": True},
                    {"value": "Closed", "bold": True},
                    {"value": "Status", "bold": True},
                ],
                ["North", 1500.5, {"value": 0.125, "format": "0.0%"}, True,
                 {"value": "2026-01-31", "type": "date",
                  "format": "yyyy-mm-dd"}, "Yes"],
                ["South", 900, {"value": -0.03, "format": "0.0%"}, False,
                 {"value": "2026-02-28", "type": "date",
                  "format": "yyyy-mm-dd"}, "No"],
                ["East", 2100, {"value": 0.4, "format": "0.0%"}, True,
                 {"value": "2026-03-31", "type": "date",
                  "format": "yyyy-mm-dd"}, "Yes"],
            ],
            "cells": {
                "A6": {"value": "Total", "bold": True, "italic": True,
                       "font_size": 12, "font_color": "1F4E78"},
                "B6": {"formula": "SUM(B2:B4)", "format": "$#,##0.00"},
            },
            "column_widths": {"A": 18, "B": 14},
            "row_heights": {"1": 24},
            "merges": ["A8:C8"],
            "freeze_panes": "A2",
            "autofilter": "A1:F4",
            "conditional_formats": [
                {"range": "B2:B4", "type": "cell_is",
                 "operator": "greaterThan", "formula": ["1000"],
                 "fill": "C6EFCE"},
                {"range": "C2:C4", "type": "color_scale"},
            ],
            "charts": [
                {"type": "bar", "title": "Sales by region", "anchor": "H2",
                 "data": "B1:B4", "categories": "A2:A4"},
                {"type": "line", "title": "Growth", "anchor": "H18",
                 "data": "C1:C4", "categories": "A2:A4"},
                {"type": "pie", "title": "Share", "anchor": "P2",
                 "data": "B2:B4", "categories": "A2:A4",
                 "titles_from_data": False},
            ],
            "validations": [
                {"range": "F2:F10", "type": "list",
                 "formula1": '"Yes,No,Maybe"'},
            ],
        },
        {"name": "Notes", "rows": [["Zürich", "Фамилия", "12,5%"]]},
    ],
}


@pytest.fixture
def workbook(tmp_path):
    spec_path = tmp_path / "spec.json"
    spec_path.write_text(json.dumps(SPEC), encoding="utf-8")
    out = tmp_path / "report.xlsx"
    proc = run("xlsx_create.py", spec_path, out)
    summary = json.loads(proc.stdout)
    assert summary["ok"] and summary["sheets"] == ["Data", "Notes"]
    return out


def test_create_features_roundtrip(workbook):
    wb = load_workbook(workbook)
    ws = wb["Data"]
    # typed values
    assert ws["B2"].value == 1500.5
    assert ws["D2"].value is True
    e2 = ws["E2"].value
    assert (e2.date() if hasattr(e2, "date") else e2) == date(2026, 1, 31)
    # formula + number formats
    assert ws["B6"].value == "=SUM(B2:B4)"
    assert ws["B6"].number_format == "$#,##0.00"
    assert ws["C2"].number_format == "0.0%"
    assert ws["E2"].number_format == "yyyy-mm-dd"
    # styling
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fgColor.rgb.endswith("DDEBF7")
    assert ws["A1"].border.left.style == "thin"
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A6"].font.italic is True and ws["A6"].font.size == 12
    # dimensions
    assert ws.column_dimensions["A"].width == 18
    assert ws.row_dimensions[1].height == 24
    # merges / freeze / autofilter
    assert "A8:C8" in [str(r) for r in ws.merged_cells.ranges]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:F4"
    # conditional formatting, charts, validation
    assert len(list(ws.conditional_formatting)) == 2
    assert len(ws._charts) == 3
    types = {type(c).__name__ for c in ws._charts}
    assert types == {"BarChart", "LineChart", "PieChart"}
    assert len(ws.data_validations.dataValidation) == 1
    # recalc flag
    assert wb.calculation.fullCalcOnLoad is True


def test_read_sheets_json_formulas(workbook, tmp_path):
    inv = json.loads(run("xlsx_read.py", workbook, "--sheets").stdout)
    names = [s["name"] for s in inv["sheets"]]
    assert names == ["Data", "Notes"]
    data_info = inv["sheets"][0]
    assert data_info["charts"] == 3
    assert "A8:C8" in data_info["merged"]
    assert data_info["freeze_panes"] == "A2"

    dump = json.loads(
        run("xlsx_read.py", workbook, "--json", "--sheet", "Data").stdout)
    assert dump["rows"][1][0] == "North"
    assert dump["rows"][1][4] == "2026-01-31T00:00:00"

    notes = json.loads(
        run("xlsx_read.py", workbook, "--json", "--sheet", "Notes").stdout)
    assert notes["rows"][0] == ["Zürich", "Фамилия", "12,5%"]

    formulas = json.loads(run("xlsx_read.py", workbook, "--formulas").stdout)
    entry = [f for f in formulas["formulas"] if f["cell"] == "B6"][0]
    assert entry["formula"] == "=SUM(B2:B4)"
    # openpyxl never computes: cached value absent on a fresh file
    assert entry["cached"] is None

    csv_out = tmp_path / "data.csv"
    run("xlsx_read.py", workbook, "--csv", "--sheet", "Notes",
        "--out", csv_out)
    text = csv_out.read_text(encoding="utf-8")
    assert "Zürich" in text and "Фамилия" in text


def test_csv_roundtrip_nonascii(tmp_path):
    src = tmp_path / "src.csv"
    with open(src, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["City", "Share", "Surname", "Active", "When"])
        w.writerow(["Zürich", "12,5%", "Фамилия", "true", "2026-05-01"])
        w.writerow(["Oslo", "7", "Ås", "false", "2026-06-01"])
    xlsx = tmp_path / "conv.xlsx"
    run("csv_to_xlsx.py", src, xlsx, "--sheet-name", "Import")

    wb = load_workbook(xlsx)
    ws = wb["Import"]
    assert ws["A2"].value == "Zürich"
    assert ws["B2"].value == "12,5%"      # decimal comma stays a string
    assert ws["C2"].value == "Фамилия"
    assert ws["D2"].value is True          # bool inferred
    assert ws["E2"].value.date() == date(2026, 5, 1)  # date inferred
    assert ws["B3"].value == 7             # int inferred
    assert ws["A1"].font.bold is True      # styled header
    assert ws.freeze_panes == "A2"

    back = tmp_path / "back.csv"
    run("xlsx_to_csv.py", xlsx, back, "--sheet", "Import")
    with open(back, newline="", encoding="utf-8") as fh:
        rows = list(csv.reader(fh))
    assert rows[1][0] == "Zürich"
    assert rows[1][2] == "Фамилия"
    assert rows[1][3] == "True"
    assert rows[1][4] == "2026-05-01"

    # encoding override
    latin = tmp_path / "latin.csv"
    run("xlsx_to_csv.py", xlsx, latin, "--sheet", "Import",
        "--encoding", "utf-8-sig")
    assert latin.read_bytes().startswith(b"\xef\xbb\xbf")


def test_edit_existing(workbook, tmp_path):
    edited = tmp_path / "edited.xlsx"
    proc = run("xlsx_edit.py", workbook, "--sheet", "Notes",
               "--out", edited,
               "--copy-sheet", "Notes:Backup",
               "--rename-sheet", "Data:Main",
               "--set", "B1=Änderung",
               "--set", "C1=99.5",
               "--set", "D1=2026-12-24",
               "--set", "E1==SUM(C1:C1)",
               "--append", '["appended", 1, false]',
               "--insert-rows", "1:1",
               "--recalc")
    result = json.loads(proc.stdout)
    assert result["ok"]

    wb = load_workbook(edited)
    assert set(wb.sheetnames) == {"Main", "Notes", "Backup"}
    ws = wb["Notes"]
    # insert-rows ran before --set per documented order, so row 1 is blank
    # and original data moved to row 2... check documented ordering:
    # structural ops run before --set, so B1 etc. were written after insert.
    assert ws["B1"].value == "Änderung"
    assert ws["C1"].value == 99.5
    assert ws["D1"].value.date() == date(2026, 12, 24)
    assert ws["E1"].value == "=SUM(C1:C1)"
    assert wb.calculation.fullCalcOnLoad is True
    # appended row present
    found = [r for r in ws.iter_rows(values_only=True)
             if r and r[0] == "appended"]
    assert found and found[0][1] == 1 and found[0][2] is False
    # copy preserved data
    assert wb["Backup"]["A1"].value == "Zürich"


def test_help_and_errors():
    for script in ["xlsx_create.py", "xlsx_read.py", "xlsx_edit.py",
                   "csv_to_xlsx.py", "xlsx_to_csv.py",
                   "xlsx_restructure.py", "xlsx_recalc.py"]:
        proc = run(script, "--help")
        assert "usage" in proc.stdout.lower()
    bad = run("xlsx_read.py", "/nonexistent.xlsx", "--sheets",
              expect_ok=False)
    assert bad.returncode != 0
    assert json.loads(bad.stderr)["ok"] is False


# ---------------------------------------------------------------------------
# Reference-aware restructuring (xlsx_restructure.py)
# ---------------------------------------------------------------------------

RESTRUCTURE_SPEC = {
    "defined_names": {"SalesRange": "'Data'!$B$2:$B$4"},
    "sheets": [
        {
            "name": "Data",
            "rows": [
                ["Region", "Sales", "Weight"],
                ["North", 100, 0.5],
                ["South", 200, 0.3],
                ["East", 300, 0.2],
                [None, None, None],
                ["Total", None, None],
            ],
            "cells": {
                "B6": {"formula": "SUM(B2:B4)"},
                "C6": {"formula": "$B$2*C2"},
                "D6": {"formula": "LOG10(B4)"},
                "E6": {"formula": "SUM(B:B)"},
                "F6": {"formula": '"row B2: "&B2'},
            },
            "merges": ["E2:E4", "A7:B7"],
            "freeze_panes": "A2",
            "autofilter": "A1:C4",
            "conditional_formats": [
                {"range": "B2:B4", "type": "cell_is",
                 "operator": "greaterThan", "formula": ["150"],
                 "fill": "C6EFCE"},
            ],
            "validations": [
                {"range": "C2:C4", "type": "list",
                 "formula1": '"0.2,0.3,0.5"'},
            ],
            "tables": [
                {"name": "SalesTbl", "range": "A1:C4"},
            ],
        },
        {
            "name": "Summary",
            "rows": [["Grand total"]],
            "cells": {
                "B1": {"formula": "SUM(Data!B2:B4)"},
                "B2": {"formula": "'Data'!$B$3"},
                "B3": {"formula": "SUM(A1:A1)"},
            },
        },
    ],
}


@pytest.fixture
def restructure_book(tmp_path):
    spec_path = tmp_path / "rspec.json"
    spec_path.write_text(json.dumps(RESTRUCTURE_SPEC), encoding="utf-8")
    out = tmp_path / "restructure.xlsx"
    run("xlsx_create.py", spec_path, out)
    return out


def test_restructure_insert_rows_shifts_everything(restructure_book):
    # merge A6:C6 gets pushed down; A1:A1 merge is before the insert point
    proc = run("xlsx_restructure.py", restructure_book,
               "--sheet", "Data", "--insert-rows", "3:2")
    report = json.loads(proc.stdout)
    assert report["ok"] and report["op"] == "insert"

    wb = load_workbook(restructure_book)
    data, summary = wb["Data"], wb["Summary"]
    # values physically moved
    assert data["A2"].value == "North"
    assert data["A5"].value == "South"      # was row 3
    assert data["A8"].value == "Total"      # was row 6
    # same-sheet formulas rewritten (range expanded across insert point)
    assert data["B8"].value == "=SUM(B2:B6)"
    # absolute ref before insert point unchanged; relative arm shifted
    assert data["C8"].value == "=$B$2*C2"
    # function names, whole-column refs, string literals untouched
    assert data["D8"].value == "=LOG10(B6)"
    assert data["E8"].value == "=SUM(B:B)"
    assert data["F8"].value == '="row B2: "&B2'
    # cross-sheet formulas on the OTHER sheet rewritten
    assert summary["B1"].value == "=SUM(Data!B2:B6)"
    assert summary["B2"].value == "='Data'!$B$5"
    # Summary-local refs not confused with Data refs
    assert summary["B3"].value == "=SUM(A1:A1)"
    # merges: E2:E4 spans the insert point -> expanded; A7:B7 -> shifted
    merged = [str(r) for r in data.merged_cells.ranges]
    assert "E2:E6" in merged and "A9:B9" in merged
    # autofilter expanded, freeze panes intact
    assert data.auto_filter.ref == "A1:C6"
    assert data.freeze_panes == "A2"
    # validation + conditional format ranges shifted
    dv = data.data_validations.dataValidation[0]
    assert str(dv.sqref) == "C2:C6"
    cf = list(data.conditional_formatting)[0]
    assert str(cf.sqref) == "B2:B6"
    # native table expanded
    assert data.tables["SalesTbl"].ref == "A1:C6"
    # defined name rewritten
    assert wb.defined_names["SalesRange"].attr_text == "'Data'!$B$2:$B$6"
    # report is honest about limits
    assert "chart anchors" in report["not_shifted"]
    assert any(f["cell"] == "B1" and f["sheet"] == "Summary"
               for f in report["formulas"])


def test_restructure_delete_rows_and_ref_errors(restructure_book):
    run("xlsx_restructure.py", restructure_book,
        "--sheet", "Data", "--delete-rows", "3")
    wb = load_workbook(restructure_book)
    data, summary = wb["Data"], wb["Summary"]
    assert data["A3"].value == "East"           # South deleted
    assert data["B5"].value == "=SUM(B2:B3)"    # range clamped
    # single-cell ref into the deleted row becomes #REF!
    assert summary["B2"].value == "='Data'!#REF!"
    assert summary["B1"].value == "=SUM(Data!B2:B3)"
    assert data.tables["SalesTbl"].ref == "A1:C3"


def test_restructure_insert_cols(restructure_book):
    proc = run("xlsx_restructure.py", restructure_book,
               "--sheet", "Data", "--insert-cols", "B:1")
    report = json.loads(proc.stdout)
    assert report["axis"] == "cols" and report["index"] == 2
    wb = load_workbook(restructure_book)
    data, summary = wb["Data"], wb["Summary"]
    assert data["C2"].value == 100              # Sales moved B->C
    assert data["C6"].value == "=SUM(C2:C4)"
    assert data["D6"].value == "=$C$2*D2"
    assert summary["B1"].value == "=SUM(Data!C2:C4)"
    assert wb.defined_names["SalesRange"].attr_text == "'Data'!$C$2:$C$4"
    merged = [str(r) for r in data.merged_cells.ranges]
    assert "F2:F4" in merged                    # merge shifted right
    assert "A7:C7" in merged                    # merge expanded across col B


# ---------------------------------------------------------------------------
# Tables, defined names, hyperlinks, notes, protection (edit + read paths)
# ---------------------------------------------------------------------------

def test_tables_create_append_list(tmp_path):
    spec = {"sheets": [{"name": "T",
                        "rows": [["Item", "Qty"], ["a", 1], ["b", 2]],
                        "tables": [{"name": "Stock", "range": "A1:B3",
                                    "style": "TableStyleLight1"}]}]}
    spec_path = tmp_path / "tspec.json"
    spec_path.write_text(json.dumps(spec), encoding="utf-8")
    book = tmp_path / "tables.xlsx"
    run("xlsx_create.py", spec_path, book)

    wb = load_workbook(book)
    tbl = wb["T"].tables["Stock"]
    assert tbl.ref == "A1:B3"
    assert tbl.tableStyleInfo.name == "TableStyleLight1"

    # --add-table + --table-append auto-extends the range
    run("xlsx_edit.py", book, "--sheet", "T",
        "--add-table", "Extra:D1:E2",
        "--table-append", 'Stock=["c", 3]')
    wb = load_workbook(book)
    ws = wb["T"]
    assert ws.tables["Stock"].ref == "A1:B4"
    assert ws["A4"].value == "c" and ws["B4"].value == 3
    assert ws.tables["Extra"].ref == "D1:E2"

    listing = json.loads(
        run("xlsx_edit.py", book, "--sheet", "T", "--list-tables").stdout)
    assert listing["tables"]["Stock"]["ref"] == "A1:B4"
    assert set(listing["tables"]) == {"Stock", "Extra"}
    # tables also appear in the read inventory
    inv = json.loads(run("xlsx_read.py", book, "--sheets").stdout)
    assert inv["sheets"][0]["tables"]["Stock"] == "A1:B4"


def test_names_hyperlinks_notes(tmp_path):
    spec = {
        "defined_names": {"Rate": "'D'!$B$1"},
        "sheets": [{"name": "D", "cells": {
            "A1": {"value": "docs",
                   "hyperlink": "https://example.com/docs"},
            "B1": {"value": 0.07, "note": "quarterly rate"},
            "C1": {"value": 1, "note": {"text": "check", "author": "QA"}},
        }}],
    }
    spec_path = tmp_path / "nspec.json"
    spec_path.write_text(json.dumps(spec), encoding="utf-8")
    book = tmp_path / "names.xlsx"
    run("xlsx_create.py", spec_path, book)

    wb = load_workbook(book)
    ws = wb["D"]
    assert ws["A1"].hyperlink.target == "https://example.com/docs"
    assert ws["B1"].comment.text == "quarterly rate"
    assert ws["C1"].comment.author == "QA"
    assert wb.defined_names["Rate"].attr_text == "'D'!$B$1"

    # edit path: add/delete names, hyperlink, note, clear note
    run("xlsx_edit.py", book, "--sheet", "D",
        "--define-name", "Extra='D'!$C$1",
        "--delete-name", "Rate",
        "--hyperlink", "D1=https://example.com/more|More",
        "--note", "D1=see more|Reviewer",
        "--clear-note", "B1")
    wb = load_workbook(book)
    ws = wb["D"]
    assert "Rate" not in wb.defined_names
    assert wb.defined_names["Extra"].attr_text == "'D'!$C$1"
    assert ws["D1"].hyperlink.target == "https://example.com/more"
    assert ws["D1"].value == "More"
    assert ws["D1"].comment.author == "Reviewer"
    assert ws["B1"].comment is None

    # read path: --notes and --names JSON output
    notes = json.loads(run("xlsx_read.py", book, "--notes").stdout)["notes"]
    coords = {(n["cell"], n["author"]) for n in notes}
    assert ("D1", "Reviewer") in coords and ("C1", "QA") in coords
    names = json.loads(run("xlsx_read.py", book, "--names").stdout)
    assert names["defined_names"] == {"Extra": "'D'!$C$1"}


def test_sheet_protection(tmp_path):
    spec = {"sheets": [{"name": "P", "rows": [["locked", "open"]],
                        "protection": {"password": "your-password",
                                       "unlock": ["B1:B1"]}}]}
    spec_path = tmp_path / "pspec.json"
    spec_path.write_text(json.dumps(spec), encoding="utf-8")
    book = tmp_path / "prot.xlsx"
    run("xlsx_create.py", spec_path, book)

    wb = load_workbook(book)
    ws = wb["P"]
    assert ws.protection.sheet is True
    assert ws.protection.password           # hash stored
    assert ws["B1"].protection.locked is False
    assert ws["A1"].protection.locked is not False
    inv = json.loads(run("xlsx_read.py", book, "--sheets").stdout)
    assert inv["sheets"][0]["protected"] is True

    # edit path on a fresh unprotected sheet
    plain = tmp_path / "plain.xlsx"
    spec_path.write_text(json.dumps(
        {"sheets": [{"name": "P", "rows": [["a", "b"]]}]}), encoding="utf-8")
    run("xlsx_create.py", spec_path, plain)
    run("xlsx_edit.py", plain, "--sheet", "P",
        "--protect", "your-password", "--unlock", "B1:B1")
    ws = load_workbook(plain)["P"]
    assert ws.protection.sheet is True and ws["B1"].protection.locked is False


# ---------------------------------------------------------------------------
# Headless recalculation (xlsx_recalc.py) — branches on soffice presence
# ---------------------------------------------------------------------------

def test_recalc_reports_json_both_ways(tmp_path):
    spec = {"sheets": [{"name": "R", "rows": [[2], [3]],
                        "cells": {"A3": {"formula": "SUM(A1:A2)"}}}]}
    spec_path = tmp_path / "cspec.json"
    spec_path.write_text(json.dumps(spec), encoding="utf-8")
    book = tmp_path / "calc.xlsx"
    run("xlsx_create.py", spec_path, book)

    # absent-soffice branch is always testable by hiding PATH
    env = dict(os.environ, LC_ALL="C", LANG="C", PATH=str(tmp_path))
    proc = subprocess.run(
        [sys.executable, str(SCRIPTS / "xlsx_recalc.py"), str(book)],
        capture_output=True, text=True, env=env, encoding="utf-8")
    assert proc.returncode == 0
    absent = json.loads(proc.stdout)
    assert absent["recalculated"] is False and "soffice" in absent["reason"]
    assert "guidance" in absent

    if not shutil.which("soffice"):
        pytest.skip("LibreOffice not installed; absent branch covered above")

    out = tmp_path / "calced.xlsx"
    proc = run("xlsx_recalc.py", book, "--out", out, "--timeout", "300")
    result = json.loads(proc.stdout)
    assert result["recalculated"] is True
    assert result["formula_cells"] == 1
    assert result["with_cached_values"] == 1
    # cached value now visible to --formulas
    formulas = json.loads(run("xlsx_read.py", out, "--formulas").stdout)
    entry = formulas["formulas"][0]
    assert entry["formula"] == "=SUM(A1:A2)" and entry["cached"] == 5
