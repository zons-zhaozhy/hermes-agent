#!/usr/bin/env python3
"""Export one sheet of an .xlsx workbook to CSV.

Dates/datetimes are written in ISO format; None becomes an empty field.
With --data-only, formula cells yield their cached results (present only
if the file was last saved by Excel/LibreOffice; openpyxl never computes).

Usage:
  xlsx_to_csv.py book.xlsx out.csv
  xlsx_to_csv.py book.xlsx out.csv --sheet Data --encoding utf-8-sig
  xlsx_to_csv.py book.xlsx out.csv --delimiter ';' --data-only
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date, datetime, time

from openpyxl import load_workbook


def to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Excel stores pure dates as midnight datetimes; emit a bare date.
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value


def main(argv=None):
    ap = argparse.ArgumentParser(description=".xlsx sheet -> CSV exporter.")
    ap.add_argument("file", help="input .xlsx path")
    ap.add_argument("output", help="output CSV path")
    ap.add_argument("--sheet", help="sheet name (default: active)")
    ap.add_argument("--encoding", default="utf-8",
                    help="CSV output encoding (default utf-8)")
    ap.add_argument("--delimiter", default=",")
    ap.add_argument("--data-only", action="store_true",
                    help="cached formula results instead of formula strings")
    args = ap.parse_args(argv)

    wb = load_workbook(args.file, data_only=args.data_only)
    ws = wb[args.sheet] if args.sheet else wb.active

    with open(args.output, "w", newline="", encoding=args.encoding) as fh:
        writer = csv.writer(fh, delimiter=args.delimiter)
        count = 0
        for row in ws.iter_rows(values_only=True):
            writer.writerow([to_text(v) for v in row])
            count += 1

    print(json.dumps({"ok": True, "output": args.output, "sheet": ws.title,
                      "rows": count}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
