#!/usr/bin/env python3
"""Read an .xlsx workbook: inventory, JSON/CSV dumps, formula listing.

Modes (pick one):
  --sheets     JSON inventory: sheet names, dimensions, row/col counts
  --json       dump one sheet's rows as a JSON array of arrays
  --csv        dump one sheet as CSV to stdout or --out
  --formulas   JSON list of formula cells {"cell", "formula", "cached"}
  --notes      JSON list of cell notes/comments across sheets
  --names      JSON map of workbook defined names

Options:
  --sheet NAME     sheet to dump (default: active sheet)
  --data-only      load cached formula RESULTS instead of formula strings.
                   Caveat: openpyxl never computes formulas; cached values
                   exist only if the file was last saved by Excel/LibreOffice.
  --encoding ENC   encoding for --csv --out files (default utf-8)
  --out PATH       write --csv output to a file instead of stdout

Usage:
  xlsx_read.py book.xlsx --sheets
  xlsx_read.py book.xlsx --json --sheet Data
  xlsx_read.py book.xlsx --csv --sheet Data --out data.csv
  xlsx_read.py book.xlsx --formulas
  xlsx_read.py book.xlsx --notes
  xlsx_read.py book.xlsx --names
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date, datetime, time

from openpyxl import load_workbook


def jsonable(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def sheet_rows(ws):
    return [[jsonable(c) for c in row] for row in ws.iter_rows(values_only=True)]


def cmd_sheets(wb):
    info = []
    for ws in wb.worksheets:
        info.append({
            "name": ws.title,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged": [str(r) for r in ws.merged_cells.ranges],
            "charts": len(getattr(ws, "_charts", [])),
            "freeze_panes": ws.freeze_panes,
            "autofilter": ws.auto_filter.ref,
            "tables": {t.displayName: t.ref for t in ws.tables.values()},
            "protected": bool(ws.protection.sheet),
        })
    names = {name: dn.attr_text for name, dn in wb.defined_names.items()}
    print(json.dumps({"sheets": info, "defined_names": names},
                     ensure_ascii=False, indent=2))


def cmd_notes(wb, sheet):
    out = []
    sheets = [sheet] if sheet else wb.sheetnames
    for name in sheets:
        for row in wb[name].iter_rows():
            for cell in row:
                if cell.comment is not None:
                    out.append({"sheet": name, "cell": cell.coordinate,
                                "text": cell.comment.text,
                                "author": cell.comment.author})
    print(json.dumps({"notes": out}, ensure_ascii=False, indent=2))


def cmd_names(wb):
    names = {name: dn.attr_text for name, dn in wb.defined_names.items()}
    print(json.dumps({"defined_names": names}, ensure_ascii=False, indent=2))


def cmd_formulas(path, sheet):
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    out = []
    sheets = [sheet] if sheet else wb_f.sheetnames
    for name in sheets:
        ws_f, ws_v = wb_f[name], wb_v[name]
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    out.append({
                        "sheet": name,
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "cached": jsonable(ws_v[cell.coordinate].value),
                    })
    print(json.dumps({"formulas": out}, ensure_ascii=False, indent=2))


def main(argv=None):
    ap = argparse.ArgumentParser(description="Read/inspect an .xlsx workbook.")
    ap.add_argument("file", help="path to .xlsx file")
    mode = ap.add_mutually_exclusive_group(required=True)
    mode.add_argument("--sheets", action="store_true")
    mode.add_argument("--json", action="store_true")
    mode.add_argument("--csv", action="store_true")
    mode.add_argument("--formulas", action="store_true")
    mode.add_argument("--notes", action="store_true")
    mode.add_argument("--names", action="store_true")
    ap.add_argument("--sheet", help="sheet name (default: active)")
    ap.add_argument("--data-only", action="store_true",
                    help="return cached formula results (see module docstring)")
    ap.add_argument("--encoding", default="utf-8")
    ap.add_argument("--out", help="output file for --csv")
    args = ap.parse_args(argv)

    if args.formulas:
        cmd_formulas(args.file, args.sheet)
        return 0

    wb = load_workbook(args.file, data_only=args.data_only)
    if args.sheets:
        cmd_sheets(wb)
        return 0
    if args.notes:
        cmd_notes(wb, args.sheet)
        return 0
    if args.names:
        cmd_names(wb)
        return 0

    ws = wb[args.sheet] if args.sheet else wb.active
    rows = sheet_rows(ws)
    if args.json:
        print(json.dumps({"sheet": ws.title, "rows": rows}, ensure_ascii=False))
    else:  # --csv
        if args.out:
            with open(args.out, "w", newline="", encoding=args.encoding) as fh:
                csv.writer(fh).writerows(
                    [["" if v is None else v for v in r] for r in rows])
            print(json.dumps({"ok": True, "out": args.out, "rows": len(rows)}))
        else:
            w = csv.writer(sys.stdout)
            for r in rows:
                w.writerow(["" if v is None else v for v in r])
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
