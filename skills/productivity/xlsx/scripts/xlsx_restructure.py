#!/usr/bin/env python3
"""Reference-aware row/column insert and delete for .xlsx workbooks.

Unlike plain openpyxl insert_rows/delete_cols (and xlsx_edit.py's thin
wrappers), this script also rewrites everything that points at the moved
cells:

  * formula references in ALL sheets, including absolute refs ($B$2),
    ranges (B2:B9), and cross-sheet refs ('My Sheet'!A1 / Data!$B$8).
    References into a deleted region become #REF!.
  * merged-cell ranges (shifted; expanded when they span the insertion
    point; removed when fully deleted)
  * autofilter range, freeze panes, data-validation ranges,
    conditional-formatting applied ranges (sqref)
  * native table (ListObject) refs on the edited sheet
  * workbook-scope defined names that point at the edited sheet
  * row heights / column widths

It prints a JSON report of every rewrite it made and lists what it
could NOT shift (chart anchors, images, conditional-format RULE
formulas). Full rules and limits: references/restructuring.md.

One structural operation per invocation:

Usage:
  xlsx_restructure.py book.xlsx --sheet Data --insert-rows 3:2
  xlsx_restructure.py book.xlsx --sheet Data --delete-rows 5
  xlsx_restructure.py book.xlsx --sheet Data --insert-cols B:1 --out new.xlsx
  xlsx_restructure.py book.xlsx --sheet Data --delete-cols 4:2
"""
from __future__ import annotations

import argparse
import json
import re
import sys

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.utils import (column_index_from_string, get_column_letter,
                            range_boundaries)

# A1-style reference, optionally sheet-qualified, optionally a range.
# Guards: not preceded by a word char/$/. (avoids ABC123 identifiers) and
# not followed by a word char or "(" (avoids function names like LOG10().
REF_RE = re.compile(
    r"(?<![\w$.:])"
    r"(?P<sheet>(?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"(?P<start>\$?[A-Za-z]{1,3}\$?[0-9]{1,7})"
    r"(?::(?P<end>\$?[A-Za-z]{1,3}\$?[0-9]{1,7}))?"
    r"(?![\w(])")
STRING_RE = re.compile(r'"(?:[^"]|"")*"')
COORD_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)([0-9]+)$")


def shift_point(v, idx, n, delete):
    """New 1-based index for a single row/col, or None if deleted."""
    if delete:
        if v < idx:
            return v
        if v >= idx + n:
            return v - n
        return None
    return v + n if v >= idx else v


def shift_span(a, b, idx, n, delete):
    """New (start, end) for an inclusive span, or None if fully deleted."""
    if delete:
        na = a if a < idx else (a - n if a >= idx + n else idx)
        nb = b if b < idx else (b - n if b >= idx + n else idx - 1)
        return None if na > nb else (na, nb)
    return (a + n if a >= idx else a, b + n if b >= idx else b)


def shift_range(rng, axis, idx, n, delete):
    """Shift an A1 range string (no sheet prefix). None = fully deleted."""
    min_col, min_row, max_col, max_row = range_boundaries(rng)
    if axis == "rows":
        span = shift_span(min_row, max_row, idx, n, delete)
        if span is None:
            return None
        min_row, max_row = span
    else:
        span = shift_span(min_col, max_col, idx, n, delete)
        if span is None:
            return None
        min_col, max_col = span
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end and ":" not in rng else f"{start}:{end}"


class RefRewriter:
    """Rewrite A1 references in formula-like text for one shift op."""

    def __init__(self, target_sheet, axis, idx, n, delete):
        self.target = target_sheet.lower()
        self.axis, self.idx, self.n, self.delete = axis, idx, n, delete

    def _shift_coord(self, coord):
        m = COORD_RE.match(coord)
        col_abs, col, row_abs, row = m.groups()
        ci, ri = column_index_from_string(col.upper()), int(row)
        if self.axis == "rows":
            ri = shift_point(ri, self.idx, self.n, self.delete)
            if ri is None:
                return None
        else:
            ci = shift_point(ci, self.idx, self.n, self.delete)
            if ci is None:
                return None
        return f"{col_abs}{get_column_letter(ci)}{row_abs}{ri}"

    def _shift_pair(self, start, end):
        """Shift a range preserving $ flags; None = collapsed to #REF!."""
        new_start = self._shift_coord(start)
        new_end = self._shift_coord(end)
        if new_start is None or new_end is None:
            # spans may survive partial deletion: clamp via span math
            s, e = COORD_RE.match(start), COORD_RE.match(end)
            if self.axis == "rows":
                span = shift_span(int(s.group(4)), int(e.group(4)),
                                  self.idx, self.n, self.delete)
                if span is None:
                    return None
                new_start = f"{s.group(1)}{s.group(2)}{s.group(3)}{span[0]}"
                new_end = f"{e.group(1)}{e.group(2)}{e.group(3)}{span[1]}"
            else:
                span = shift_span(column_index_from_string(s.group(2).upper()),
                                  column_index_from_string(e.group(2).upper()),
                                  self.idx, self.n, self.delete)
                if span is None:
                    return None
                new_start = (f"{s.group(1)}{get_column_letter(span[0])}"
                             f"{s.group(3)}{s.group(4)}")
                new_end = (f"{e.group(1)}{get_column_letter(span[1])}"
                           f"{e.group(3)}{e.group(4)}")
        return new_start, new_end

    def _sub(self, match, home_sheet):
        prefix = match.group("sheet") or ""
        if prefix:
            name = prefix[:-1]
            if name.startswith("'"):
                name = name[1:-1].replace("''", "'")
            ref_sheet = name
        else:
            ref_sheet = home_sheet
        if ref_sheet.lower() != self.target:
            return match.group(0)
        start, end = match.group("start"), match.group("end")
        if end is None:
            new = self._shift_coord(start)
            return prefix + ("#REF!" if new is None else new)
        pair = self._shift_pair(start, end)
        return (prefix + "#REF!" if pair is None
                else f"{prefix}{pair[0]}:{pair[1]}")

    def rewrite(self, text, home_sheet):
        """Rewrite refs outside quoted string literals. Returns new text."""
        out, pos = [], 0
        for lit in STRING_RE.finditer(text):
            out.append(REF_RE.sub(lambda m: self._sub(m, home_sheet),
                                  text[pos:lit.start()]))
            out.append(lit.group(0))
            pos = lit.end()
        out.append(REF_RE.sub(lambda m: self._sub(m, home_sheet), text[pos:]))
        return "".join(out)


def shift_dimensions(dims, idx, n, delete, is_row):
    """Rebuild a row/column dimensions map with shifted keys."""
    items = list(dims.items())
    saved = {}
    for key, dim in items:
        pos = key if is_row else column_index_from_string(key)
        new = shift_point(pos, idx, n, delete)
        if new is not None and new != pos:
            saved[new if is_row else get_column_letter(new)] = dim
            del dims[key]
    for key, dim in saved.items():
        if is_row:
            dim.index = key
        else:
            dim.index = column_index_from_string(key)
        dims[key] = dim
    return len(saved)


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Insert/delete rows or columns AND rewrite formula "
                    "references, merges, filters, validations, tables, and "
                    "defined names to match.",
        epilog="Cannot shift: chart anchors, images, conditional-format "
               "rule formulas. See references/restructuring.md.")
    ap.add_argument("file", help="path to .xlsx file")
    ap.add_argument("--sheet", help="target sheet (default: active)")
    ap.add_argument("--out", help="output path (default: edit in place)")
    op = ap.add_mutually_exclusive_group(required=True)
    op.add_argument("--insert-rows", metavar="IDX[:N]")
    op.add_argument("--delete-rows", metavar="IDX[:N]")
    op.add_argument("--insert-cols", metavar="COL[:N]",
                    help="COL is a letter (B) or 1-based number")
    op.add_argument("--delete-cols", metavar="COL[:N]")
    args = ap.parse_args(argv)

    raw = (args.insert_rows or args.delete_rows
           or args.insert_cols or args.delete_cols)
    idx_s, _, n_s = raw.partition(":")
    n = int(n_s) if n_s else 1
    axis = "rows" if (args.insert_rows or args.delete_rows) else "cols"
    delete = bool(args.delete_rows or args.delete_cols)
    if axis == "cols" and idx_s.isalpha():
        idx = column_index_from_string(idx_s.upper())
    else:
        idx = int(idx_s)

    wb = load_workbook(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active
    rewriter = RefRewriter(ws.title, axis, idx, n, delete)
    report = {"ok": True, "sheet": ws.title, "axis": axis,
              "op": "delete" if delete else "insert", "index": idx, "count": n,
              "formulas": [], "merges": [], "tables": {}, "defined_names": {},
              "validations": [], "conditional_formats": [],
              "not_shifted": ["chart anchors", "images",
                              "conditional-format rule formulas"]}

    # 1. capture merge ranges (openpyxl does not move them), then unmerge
    old_merges = [str(r) for r in list(ws.merged_cells.ranges)]
    for rng in old_merges:
        ws.unmerge_cells(rng)

    # 2. structural move of cell values/styles/comments
    getattr(ws, f"{report['op']}_{axis}")(idx, n)

    # 3. formulas everywhere
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    new = rewriter.rewrite(cell.value, sheet.title)
                    if new != cell.value:
                        report["formulas"].append(
                            {"sheet": sheet.title, "cell": cell.coordinate,
                             "from": cell.value, "to": new})
                        cell.value = new

    # 4. merges back, shifted
    for rng in old_merges:
        new = shift_range(rng, axis, idx, n, delete)
        if new is None:
            report["merges"].append({"from": rng, "to": None})
        else:
            ws.merge_cells(new)
            if new != rng:
                report["merges"].append({"from": rng, "to": new})

    # 5. autofilter + freeze panes
    if ws.auto_filter.ref:
        new = shift_range(ws.auto_filter.ref, axis, idx, n, delete)
        if new != ws.auto_filter.ref:
            report["autofilter"] = {"from": ws.auto_filter.ref, "to": new}
            ws.auto_filter.ref = new
    if ws.freeze_panes:
        m = COORD_RE.match(ws.freeze_panes)
        ci = column_index_from_string(m.group(2).upper())
        ri = int(m.group(4))
        if axis == "rows":
            ri = shift_point(ri, idx, n, delete) or max(idx, 2)
        else:
            ci = shift_point(ci, idx, n, delete) or max(idx, 2)
        new = f"{get_column_letter(ci)}{ri}"
        if new != ws.freeze_panes:
            report["freeze_panes"] = {"from": ws.freeze_panes, "to": new}
            ws.freeze_panes = new

    # 6. data validations + conditional formatting applied ranges
    for dv in ws.data_validations.dataValidation:
        old = str(dv.sqref)
        parts = [shift_range(p, axis, idx, n, delete) for p in old.split()]
        parts = [p for p in parts if p]
        if parts and " ".join(parts) != old:
            dv.sqref = " ".join(parts)
            report["validations"].append({"from": old, "to": str(dv.sqref)})
    new_cf = ConditionalFormattingList()
    for cf in ws.conditional_formatting:
        old = str(cf.sqref)
        parts = [shift_range(p, axis, idx, n, delete) for p in old.split()]
        parts = [p for p in parts if p]
        if not parts:
            report["conditional_formats"].append({"from": old, "to": None})
            continue
        new = " ".join(parts)
        for rule in cf.rules:
            new_cf.add(new, rule)
        if new != old:
            report["conditional_formats"].append({"from": old, "to": new})
    ws.conditional_formatting = new_cf

    # 7. native tables on the edited sheet
    for table in ws.tables.values():
        new = shift_range(table.ref, axis, idx, n, delete)
        if new and new != table.ref:
            report["tables"][table.displayName] = {"from": table.ref,
                                                   "to": new}
            table.ref = new

    # 8. workbook-scope defined names
    for name, dn in wb.defined_names.items():
        if dn.attr_text and "!" in dn.attr_text:
            new = rewriter.rewrite(dn.attr_text, ws.title)
            if new != dn.attr_text:
                report["defined_names"][name] = {"from": dn.attr_text,
                                                 "to": new}
                dn.attr_text = new

    # 9. row heights / column widths
    if axis == "rows":
        shift_dimensions(ws.row_dimensions, idx, n, delete, is_row=True)
    else:
        shift_dimensions(ws.column_dimensions, idx, n, delete, is_row=False)

    out = args.out or args.file
    wb.save(out)
    report["output"] = out
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
