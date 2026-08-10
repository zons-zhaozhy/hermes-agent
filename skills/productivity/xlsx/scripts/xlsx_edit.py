#!/usr/bin/env python3
"""Edit an existing .xlsx workbook in place (or to --out).

Operations (repeatable where noted, applied in the order listed below):
  --rename-sheet OLD:NEW        rename a sheet
  --copy-sheet SRC:NEW          duplicate a sheet under a new name
  --insert-rows IDX[:N]         insert N rows before row IDX (default N=1)
  --delete-rows IDX[:N]         delete N rows starting at row IDX
  --insert-cols IDX[:N]         insert N columns before column IDX (number)
  --delete-cols IDX[:N]         delete N columns starting at column IDX
  --set CELL=VALUE              repeatable; type-inferred (int, float, bool,
                                ISO date, else string). '=...' sets a formula.
  --append ROWJSON              repeatable; JSON array appended as a row
  --add-table NAME:RANGE[:STYLE]  create a native Excel table (ListObject)
  --table-append NAME=ROWJSON   append a row inside a table, auto-extending
                                the table's range (repeatable)
  --list-tables                 print tables on the target sheet and exit
  --define-name NAME=REF        workbook-scope defined name, e.g.
                                "Rates='Data'!$B$2:$B$9" (repeatable)
  --delete-name NAME            remove a defined name (repeatable)
  --hyperlink CELL=URL[|TEXT]   set a hyperlink (optional display text)
  --note CELL=TEXT[|AUTHOR]     set a cell note/comment (repeatable)
  --clear-note CELL             remove a cell note (repeatable)
  --protect [PASSWORD]          enable sheet protection; combine with
                                --unlock RANGE to leave ranges editable.
                                NOT security: trivially strippable (see
                                SKILL.md Pitfalls).
  --recalc                      set fullCalcOnLoad so Excel/LibreOffice
                                recomputes all formulas on next open

WARNING: openpyxl does NOT shift merged-cell ranges, chart anchors, or
formula references when rows/columns are inserted or deleted. Verify any
sheet containing merges or formulas after structural edits — or use
xlsx_restructure.py, which rewrites references for you.

Usage:
  xlsx_edit.py book.xlsx --sheet Data --set B2=42 --set C2=2026-01-01 \
      --set "D2==SUM(B2:C2)" --recalc
  xlsx_edit.py book.xlsx --sheet Data --append '["Widget", 9.99, true]'
  xlsx_edit.py book.xlsx --copy-sheet Data:Backup --rename-sheet Data:Main
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo


def infer(text):
    if text.startswith("="):
        return text  # formula
    low = text.lower()
    if low in ("true", "false"):
        return low == "true"
    for caster in (int, float):
        try:
            return caster(text)
        except ValueError:
            pass
    for parser in (date.fromisoformat, datetime.fromisoformat):
        try:
            return parser(text)
        except ValueError:
            pass
    return text


def parse_idx(arg):
    if ":" in arg:
        idx, n = arg.split(":", 1)
        return int(idx), int(n)
    return int(arg), 1


def add_table(ws, spec):
    parts = spec.split(":")
    if len(parts) < 3:
        raise ValueError("--add-table needs NAME:RANGE like Sales:A1:C9")
    name = parts[0]
    rng = ":".join(parts[1:3])
    style = parts[3] if len(parts) > 3 else "TableStyleMedium9"
    table = Table(displayName=name, ref=rng)
    table.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(table)


def table_append(ws, name, row_values):
    table = ws.tables[name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    new_row = max_row + 1
    for offset, value in enumerate(row_values):
        ws.cell(row=new_row, column=min_col + offset, value=value)
    table.ref = (f"{get_column_letter(min_col)}{min_row}:"
                 f"{get_column_letter(max_col)}{new_row}")


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Edit an existing .xlsx workbook.",
        epilog="Plain insert/delete does not shift merges/formula refs — "
               "use xlsx_restructure.py for reference-aware moves.")
    ap.add_argument("file", help="path to .xlsx file")
    ap.add_argument("--sheet", help="target sheet (default: active)")
    ap.add_argument("--out", help="output path (default: edit in place)")
    ap.add_argument("--rename-sheet", action="append", default=[],
                    metavar="OLD:NEW")
    ap.add_argument("--copy-sheet", action="append", default=[],
                    metavar="SRC:NEW")
    ap.add_argument("--insert-rows", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--delete-rows", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--insert-cols", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--delete-cols", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--set", action="append", default=[], metavar="CELL=VALUE")
    ap.add_argument("--append", action="append", default=[], metavar="ROWJSON")
    ap.add_argument("--add-table", action="append", default=[],
                    metavar="NAME:RANGE[:STYLE]")
    ap.add_argument("--table-append", action="append", default=[],
                    metavar="NAME=ROWJSON")
    ap.add_argument("--list-tables", action="store_true",
                    help="print tables on the target sheet and exit")
    ap.add_argument("--define-name", action="append", default=[],
                    metavar="NAME=REF")
    ap.add_argument("--delete-name", action="append", default=[],
                    metavar="NAME")
    ap.add_argument("--hyperlink", action="append", default=[],
                    metavar="CELL=URL[|TEXT]")
    ap.add_argument("--note", action="append", default=[],
                    metavar="CELL=TEXT[|AUTHOR]")
    ap.add_argument("--clear-note", action="append", default=[],
                    metavar="CELL")
    ap.add_argument("--protect", nargs="?", const="", metavar="PASSWORD",
                    help="protect the target sheet (integrity signal only, "
                    "NOT security)")
    ap.add_argument("--unlock", action="append", default=[], metavar="RANGE",
                    help="cell range left editable under --protect")
    ap.add_argument("--recalc", action="store_true",
                    help="force full recalculation when the file is opened")
    args = ap.parse_args(argv)

    wb = load_workbook(args.file)
    changes = []

    for pair in args.rename_sheet:
        old, new = pair.split(":", 1)
        wb[old].title = new
        changes.append(f"rename {old}->{new}")
    for pair in args.copy_sheet:
        src, new = pair.split(":", 1)
        copy = wb.copy_worksheet(wb[src])
        copy.title = new
        changes.append(f"copy {src}->{new}")

    ws = wb[args.sheet] if args.sheet else wb.active

    if args.list_tables:
        print(json.dumps({"ok": True, "sheet": ws.title,
                          "tables": {t.displayName: {
                              "ref": t.ref,
                              "style": t.tableStyleInfo.name
                              if t.tableStyleInfo else None}
                              for t in ws.tables.values()}},
                         ensure_ascii=False))
        return 0

    for arg in args.insert_rows:
        idx, n = parse_idx(arg)
        ws.insert_rows(idx, n)
        changes.append(f"insert_rows {idx}x{n}")
    for arg in args.delete_rows:
        idx, n = parse_idx(arg)
        ws.delete_rows(idx, n)
        changes.append(f"delete_rows {idx}x{n}")
    for arg in args.insert_cols:
        idx, n = parse_idx(arg)
        ws.insert_cols(idx, n)
        changes.append(f"insert_cols {idx}x{n}")
    for arg in args.delete_cols:
        idx, n = parse_idx(arg)
        ws.delete_cols(idx, n)
        changes.append(f"delete_cols {idx}x{n}")

    for assignment in args.set:
        coord, raw = assignment.split("=", 1)
        ws[coord] = infer(raw)
        changes.append(f"set {coord}")
    for row_json in args.append:
        ws.append(json.loads(row_json))
        changes.append(f"append row {ws.max_row}")

    for spec in args.add_table:
        add_table(ws, spec)
        changes.append(f"add_table {spec.split(':')[0]}")
    for spec in args.table_append:
        name, row_json = spec.split("=", 1)
        table_append(ws, name, json.loads(row_json))
        changes.append(f"table_append {name} -> {ws.tables[name].ref}")

    for spec in args.define_name:
        name, ref = spec.split("=", 1)
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
        changes.append(f"define_name {name}")
    for name in args.delete_name:
        del wb.defined_names[name]
        changes.append(f"delete_name {name}")

    for spec in args.hyperlink:
        coord, rest = spec.split("=", 1)
        url, _, text = rest.partition("|")
        cell = ws[coord]
        cell.hyperlink = url
        cell.value = text or (cell.value if cell.value is not None else url)
        cell.style = "Hyperlink"
        changes.append(f"hyperlink {coord}")
    for spec in args.note:
        coord, rest = spec.split("=", 1)
        text, _, author = rest.partition("|")
        ws[coord].comment = Comment(text, author or "xlsx-skill")
        changes.append(f"note {coord}")
    for coord in args.clear_note:
        ws[coord].comment = None
        changes.append(f"clear_note {coord}")

    if args.protect is not None:
        for rng in args.unlock:
            for row in ws[rng]:
                for cell in row:
                    cell.protection = Protection(locked=False)
        if args.protect:
            ws.protection.password = args.protect
        ws.protection.sheet = True
        changes.append(f"protect {ws.title}"
                       + (f" (unlocked {len(args.unlock)} ranges)"
                          if args.unlock else ""))

    if args.recalc:
        wb.calculation.fullCalcOnLoad = True
        changes.append("fullCalcOnLoad")

    out = args.out or args.file
    wb.save(out)
    print(json.dumps({"ok": True, "output": out, "sheet": ws.title,
                      "changes": changes}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
