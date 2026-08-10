#!/usr/bin/env python3
"""Create an .xlsx workbook from a JSON spec.

Spec (JSON object):
  {
    "full_calc_on_load": true,          # force recalculation on open (optional)
    "defined_names": {"Rates": "'Data'!$B$2:$B$4"},   # workbook scope
    "sheets": [
      {
        "name": "Data",
        "rows": [["Header", 1, true], ...],   # scalars or cell objects (see below)
        "cells": {"A1": {"value": 5, "format": "0.00%"}},  # sparse overrides
        "column_widths": {"A": 22, "B": 12},
        "row_heights": {"1": 24},
        "merges": ["A1:C1"],
        "freeze_panes": "A2",
        "autofilter": "A1:C10",
        "conditional_formats": [
          {"range": "B2:B9", "type": "cell_is", "operator": "greaterThan",
           "formula": ["100"], "fill": "FFC7CE"},
          {"range": "C2:C9", "type": "color_scale"}
        ],
        "charts": [
          {"type": "bar", "title": "Sales", "anchor": "F2",
           "data": "B1:B5", "categories": "A2:A5"}
        ],
        "validations": [
          {"range": "D2:D9", "type": "list", "formula1": "\"Yes,No,Maybe\""}
        ],
        "tables": [
          {"name": "Sales", "range": "A1:C4",
           "style": "TableStyleMedium9"}        # native Excel table
        ],
        "protection": {"password": "your-password",   # NOT security --
                       "unlock": ["B2:B9"]}           # see SKILL.md Pitfalls
      }
    ]
  }

Cell object keys (all optional except value/formula):
  value          scalar; JSON true/false -> bool, numbers stay numeric
  type           "date" or "datetime" -> value parsed from ISO string
  formula        e.g. "=SUM(A2:A9)" (leading '=' optional)
  hyperlink      URL; value becomes the display text
  note           cell note text (or {"text": ..., "author": ...})
  format         Excel number format, e.g. "$#,##0.00", "0.0%", "yyyy-mm-dd"
  bold, italic   booleans
  font_size      points
  font_color     hex RGB like "FF0000"
  fill           solid fill hex RGB like "DDEBF7"
  border         "thin" | "medium" | "thick" (all four sides)
  align          "left" | "center" | "right"
  valign         "top" | "center" | "bottom"
  wrap           boolean (wrap text)

Usage:
  xlsx_create.py spec.json out.xlsx
  xlsx_create.py - out.xlsx   (spec on stdin)

Prints a JSON summary to stdout; exits non-zero on failure.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                             Protection, Side)
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


def parse_typed(value, type_hint=None):
    if type_hint == "date" and isinstance(value, str):
        return date.fromisoformat(value)
    if type_hint == "datetime" and isinstance(value, str):
        return datetime.fromisoformat(value)
    return value


def apply_cell(ws, coord, spec):
    cell = ws[coord]
    if isinstance(spec, dict):
        if "formula" in spec:
            f = spec["formula"]
            cell.value = f if f.startswith("=") else "=" + f
        elif "value" in spec:
            cell.value = parse_typed(spec["value"], spec.get("type"))
        if "hyperlink" in spec:
            cell.hyperlink = spec["hyperlink"]
            if cell.value is None:
                cell.value = spec["hyperlink"]
            cell.style = "Hyperlink"
        if "note" in spec:
            note = spec["note"]
            if isinstance(note, dict):
                cell.comment = Comment(note.get("text", ""),
                                       note.get("author", "xlsx-skill"))
            else:
                cell.comment = Comment(str(note), "xlsx-skill")
        if "format" in spec:
            cell.number_format = spec["format"]
        font_kw = {}
        if spec.get("bold"):
            font_kw["bold"] = True
        if spec.get("italic"):
            font_kw["italic"] = True
        if "font_size" in spec:
            font_kw["size"] = spec["font_size"]
        if "font_color" in spec:
            font_kw["color"] = spec["font_color"]
        if font_kw:
            cell.font = Font(**font_kw)
        if "fill" in spec:
            cell.fill = PatternFill("solid", fgColor=spec["fill"])
        if "border" in spec:
            side = Side(style=spec["border"])
            cell.border = Border(left=side, right=side, top=side, bottom=side)
        align_kw = {}
        if "align" in spec:
            align_kw["horizontal"] = spec["align"]
        if "valign" in spec:
            align_kw["vertical"] = spec["valign"]
        if spec.get("wrap"):
            align_kw["wrap_text"] = True
        if align_kw:
            cell.alignment = Alignment(**align_kw)
    else:
        cell.value = spec


def ref_from_range(ws, rng):
    min_col, min_row, max_col, max_row = range_boundaries(rng)
    return Reference(ws, min_col=min_col, min_row=min_row,
                     max_col=max_col, max_row=max_row)


def add_chart(ws, spec):
    kind = spec.get("type", "bar")
    chart = {"bar": BarChart, "line": LineChart, "pie": PieChart}[kind]()
    if "title" in spec:
        chart.title = spec["title"]
    data = ref_from_range(ws, spec["data"])
    chart.add_data(data, titles_from_data=spec.get("titles_from_data", True))
    if "categories" in spec:
        chart.set_categories(ref_from_range(ws, spec["categories"]))
    ws.add_chart(chart, spec.get("anchor", "H2"))


def add_conditional(ws, spec):
    rng = spec["range"]
    kind = spec.get("type", "cell_is")
    if kind == "color_scale":
        rule = ColorScaleRule(
            start_type="min", start_color=spec.get("start_color", "FFF8696B"),
            end_type="max", end_color=spec.get("end_color", "FF63BE7B"))
    else:
        fill = PatternFill("solid", fgColor=spec.get("fill", "FFC7CE"))
        rule = CellIsRule(operator=spec.get("operator", "greaterThan"),
                          formula=spec.get("formula", ["0"]), fill=fill)
    ws.conditional_formatting.add(rng, rule)


def build_sheet(ws, spec):
    for row in spec.get("rows", []):
        values, styled = [], []
        for item in row:
            if isinstance(item, dict):
                values.append(None)
                styled.append(item)
            else:
                values.append(item)
                styled.append(None)
        ws.append(values)
        r = ws.max_row
        for idx, item in enumerate(styled, start=1):
            if item is not None:
                apply_cell(ws, ws.cell(row=r, column=idx).coordinate, item)
    for coord, cell_spec in spec.get("cells", {}).items():
        apply_cell(ws, coord, cell_spec)
    for col, width in spec.get("column_widths", {}).items():
        ws.column_dimensions[col].width = width
    for row, height in spec.get("row_heights", {}).items():
        ws.row_dimensions[int(row)].height = height
    for rng in spec.get("merges", []):
        ws.merge_cells(rng)
    if spec.get("freeze_panes"):
        ws.freeze_panes = spec["freeze_panes"]
    if spec.get("autofilter"):
        ws.auto_filter.ref = spec["autofilter"]
    for cf in spec.get("conditional_formats", []):
        add_conditional(ws, cf)
    for ch in spec.get("charts", []):
        add_chart(ws, ch)
    for dv_spec in spec.get("validations", []):
        dv = DataValidation(type=dv_spec.get("type", "list"),
                            formula1=dv_spec["formula1"],
                            allow_blank=dv_spec.get("allow_blank", True))
        dv.add(dv_spec["range"])
        ws.add_data_validation(dv)
    for t_spec in spec.get("tables", []):
        table = Table(displayName=t_spec["name"], ref=t_spec["range"])
        table.tableStyleInfo = TableStyleInfo(
            name=t_spec.get("style", "TableStyleMedium9"),
            showRowStripes=t_spec.get("row_stripes", True),
            showColumnStripes=t_spec.get("column_stripes", False))
        ws.add_table(table)
    prot = spec.get("protection")
    if prot:
        for rng in prot.get("unlock", []):
            for row in ws[rng]:
                for cell in row:
                    cell.protection = Protection(locked=False)
        if prot.get("password"):
            ws.protection.password = prot["password"]
        ws.protection.sheet = True


def main(argv=None):
    ap = argparse.ArgumentParser(description="Create .xlsx from a JSON spec.")
    ap.add_argument("spec", help="path to JSON spec, or '-' for stdin")
    ap.add_argument("output", help="output .xlsx path")
    args = ap.parse_args(argv)

    if args.spec == "-":
        spec = json.load(sys.stdin)
    else:
        with open(args.spec, encoding="utf-8") as fh:
            spec = json.load(fh)

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_spec in spec.get("sheets", []):
        ws = wb.create_sheet(sheet_spec.get("name", "Sheet1"))
        build_sheet(ws, sheet_spec)
    for name, ref in spec.get("defined_names", {}).items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
    if spec.get("full_calc_on_load"):
        wb.calculation.fullCalcOnLoad = True
    wb.save(args.output)
    print(json.dumps({"ok": True, "output": args.output,
                      "sheets": wb.sheetnames}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
