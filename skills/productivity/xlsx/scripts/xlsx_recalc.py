#!/usr/bin/env python3
"""Recalculate a workbook's formulas headlessly with LibreOffice.

openpyxl never computes formulas. This script shells out to `soffice`
(LibreOffice) to open the workbook, recalculate, and re-save it, so
cached formula results become available to `xlsx_read.py --data-only`
and `--formulas`.

Behavior:
  * soffice on PATH: converts the file to .xlsx in a temp dir (which
    recalculates all formulas) and replaces the original (or writes
    --out). Prints {"recalculated": true, ...} and exits 0.
  * soffice absent: prints {"recalculated": false, "reason": ...} with
    installation guidance and STILL exits 0 — callers can branch on the
    JSON instead of the exit code.

Note: LibreOffice recalculates .xlsx on load per its default
calculation settings; conversion re-saves with fresh cached values.

Usage:
  xlsx_recalc.py book.xlsx
  xlsx_recalc.py book.xlsx --out recalced.xlsx
  xlsx_recalc.py book.xlsx --timeout 120
"""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path


def count_cached(path):
    """Number of formula cells with a cached value present."""
    from openpyxl import load_workbook
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    formulas = cached = 0
    for name in wb_f.sheetnames:
        ws_f, ws_v = wb_f[name], wb_v[name]
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    if ws_v[cell.coordinate].value is not None:
                        cached += 1
    return formulas, cached


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Recalculate .xlsx formulas headlessly via LibreOffice.")
    ap.add_argument("file", help="path to .xlsx file")
    ap.add_argument("--out", help="output path (default: replace input)")
    ap.add_argument("--timeout", type=int, default=180,
                    help="seconds to wait for soffice (default 180)")
    args = ap.parse_args(argv)

    src = Path(args.file).resolve()
    if not src.exists():
        print(json.dumps({"ok": False, "error": f"no such file: {src}"}),
              file=sys.stderr)
        return 1

    soffice = shutil.which("soffice")
    if not soffice:
        print(json.dumps({
            "ok": True, "recalculated": False,
            "reason": "LibreOffice (soffice) not found on PATH",
            "guidance": "Install LibreOffice (e.g. `apt install "
                        "libreoffice-calc` or `brew install --cask "
                        "libreoffice`), or open the file in Excel/"
                        "LibreOffice once and re-save it.",
        }, ensure_ascii=False))
        return 0

    with tempfile.TemporaryDirectory() as tmp:
        proc = subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx:Calc "
             "MS Excel 2007 XML", "--outdir", tmp, str(src)],
            capture_output=True, text=True, encoding="utf-8",
            timeout=args.timeout,
            env={"HOME": tmp, "PATH": Path(soffice).parent.as_posix()
                 + ":/usr/bin:/bin"})
        produced = Path(tmp) / (src.stem + ".xlsx")
        if proc.returncode != 0 or not produced.exists():
            print(json.dumps({"ok": False,
                              "error": "soffice conversion failed",
                              "stderr": proc.stderr.strip()[-500:]}),
                  file=sys.stderr)
            return 1
        formulas, cached = count_cached(produced)
        dest = Path(args.out).resolve() if args.out else src
        shutil.copyfile(produced, dest)

    print(json.dumps({
        "ok": True, "recalculated": True, "output": str(dest),
        "formula_cells": formulas, "with_cached_values": cached,
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
