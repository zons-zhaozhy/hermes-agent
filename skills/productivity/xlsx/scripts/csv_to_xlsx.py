#!/usr/bin/env python3
"""Convert a CSV file to a styled .xlsx workbook with type inference.

Type inference per cell (disable with --no-infer):
  int, float, bool ("true"/"false", case-insensitive), ISO date
  (YYYY-MM-DD) and ISO datetime; everything else stays a string.

Styling applied by default (disable with --plain):
  bold header row with a light fill, frozen top row, autofilter over the
  data range, and column widths sized to the longest cell (capped at 60).

Usage:
  csv_to_xlsx.py data.csv out.xlsx
  csv_to_xlsx.py data.csv out.xlsx --sheet-name Import --encoding cp1252
  csv_to_xlsx.py data.csv out.xlsx --delimiter ';' --no-infer
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_COL_WIDTH = 60
COL_PADDING = 2
DEFAULT_COL_WIDTH = 8


def infer(text):
    if text == "":
        return None
    low = text.lower()
    if low in ("true", "false"):
        return low == "true"
    for caster in (int, float):
        try:
            return caster(text)
        except ValueError:
            pass
    for parser in (date.fromisoformat, datetime.fromisoformat):
        try:
            return parser(text)
        except ValueError:
            pass
    return text


def main(argv=None):
    ap = argparse.ArgumentParser(description="CSV -> styled .xlsx converter.")
    ap.add_argument("csv_file", help="input CSV path")
    ap.add_argument("output", help="output .xlsx path")
    ap.add_argument("--sheet-name", default="Sheet1")
    ap.add_argument("--encoding", default="utf-8",
                    help="CSV file encoding (default utf-8)")
    ap.add_argument("--delimiter", default=",")
    ap.add_argument("--no-infer", action="store_true",
                    help="keep every cell as a string")
    ap.add_argument("--plain", action="store_true",
                    help="skip header styling / freeze / autofilter")
    args = ap.parse_args(argv)

    with open(args.csv_file, newline="", encoding=args.encoding) as fh:
        rows = list(csv.reader(fh, delimiter=args.delimiter))

    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet_name
    for i, row in enumerate(rows):
        if args.no_infer or i == 0:
            ws.append(row)
        else:
            ws.append([infer(cell) for cell in row])

    if rows and not args.plain:
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDEBF7")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx in range(1, ws.max_column + 1):
            longest = max((len(str(r[col_idx - 1])) for r in rows
                           if len(r) >= col_idx), default=DEFAULT_COL_WIDTH)
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                min(longest + COL_PADDING, MAX_COL_WIDTH)

    wb.save(args.output)
    print(json.dumps({"ok": True, "output": args.output,
                      "rows": len(rows)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
